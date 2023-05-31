from konlpy.tag import Okt
from pykospacing import Spacing
import  openpyxl  as  op  #openpyxl 모듈 import
import pandas as pd
import gpt

excel_info_0 = pd.read_excel('etestnew2.xlsx', usecols=[0])
excel_info_1 = pd.read_excel('etestnew2.xlsx', usecols=[1])
excel_info_2 = pd.read_excel('etestnew2.xlsx', usecols=[2])
excel_info_3 = pd.read_excel('etestnew2.xlsx', usecols=[3])
excel_info_4 = pd.read_excel('etestnew2.xlsx', usecols=[4])
excel_info_5 = pd.read_excel('etestnew2.xlsx', usecols=[5])
excel_info_data = pd.read_excel('etestnew2.xlsx', usecols=[6])
excel_info_7 = pd.read_excel('etestnew2.xlsx', usecols=[7])
print(type(excel_info_0))    # 형식은 pandas.core.frame.DataFrame
'''
DataFrame은 행과 열을 가진 테이블 형태의 데이터를 다루는 데에 매우 효과적이며, 많은 데이터 분석 작업에서 중요한 역할을 한다
'''

excel_info_0 = excel_info_0.values.tolist()
excel_info_1 = excel_info_1.values.tolist()
excel_info_2 = excel_info_2.values.tolist()
excel_info_3 = excel_info_3.values.tolist()
excel_info_4 = excel_info_4.values.tolist()
excel_info_5 = excel_info_5.values.tolist()
excel_info_data = excel_info_data.values.tolist()
excel_info_7 = excel_info_7.values.tolist()
print(type(excel_info_0))    # 형식은 list

okt = Okt()
spacing = Spacing()

stopword_list = []    # 불용어 리스트
full_sent_list = []
full_sent = ''    # 잘린거 다시 합치기

'''
f = open('custumStopWords.txt', 'r', encoding="UTF8")
for line in f:
    stopword_list.append(line.rstrip('\n'))
f.close()
'''
def clean(full_sent):    # 괄호안에 @지우기
    stack = []
    new_full_sent = []
    for word in full_sent:
        if stack:
            if word != "@":
                new_full_sent.append(word)
            if word == ")":
                stack.pop()
        else:
            if word == "(":
                stack.append(word)
            new_full_sent.append(word)

    return str("".join(new_full_sent))

def cutting_text(token_text):    # 특정조건에 맞게 자르기: 특정단어 다음에 특정단어 나오는거
    for idx, word in enumerate(token_text):
        '''
        try:
            word3 = int(token_text[idx+1])
            if (word == "*" and token_text[idx+1] == word3  and token_text[idx+2] != "mg"
                    and token_text[idx+2] != "mcg" and token_text[idx+2] != "IU"):
                token_text[idx] = ''
                token_text[idx+1] = ''
        except:
            pass
        '''
        '''
        try:    # @(10g) 이런형식 지우기
            if "." in word:
                word = float(word)
            else:
                word = int(word)
            
            if (token_text[idx-1] == '@(' and (token_text[idx+1] == "g" or token_text[idx+1] == "ml" or token_text[idx+1] == "mL") 
                    and token_text[idx+2] == ")" and token_text[idx+3] == "@"):
                token_text[idx-1] = ''
                token_text[idx] = ''
                token_text[idx+1] = ''
                token_text[idx+2] = ''
        except:
            pass
        '''
        try:
            if word == "@":
                for time in range(100, 0, -1):
                    time_word = str(time)+"회"
                    if token_text[idx+1] == time_word:
                        token_text[idx+1] = ''
        except:
            pass
    return token_text

def cutting_persent(token_text):
    try:    # @OOO%@ 모양 지우기
        for idx, word in enumerate(token_text):
            if "%" in word and token_text[idx-1] == "*" and token_text[idx+1] == "*":
                print(word)
                token_text[idx] = ''
                token_text[idx+1] = ''
    except:
        pass    
    '''
    try:    # @OO,OOO%@ 모양 지우기
        for idx, word in enumerate(token_text):
            if "%" in word and token_text[idx-1] == ","and token_text[idx-3] == "*" and token_text[idx+1] == "*":
             token_text[idx-2] = ''
             token_text[idx-1] = ''
             token_text[idx] = ''
             token_text[idx+1] = ''
    except:
        pass
    '''
    return token_text

def last_int_del(token_text):    # 문장 마지막에 %가 지워져서 정수만 남아있는거 지우기
    word = token_text[-1]
    try:
        word = int(word)
        token_text = token_text[:-1]
    except:
        pass
    return token_text

def last_special_del(full_sent):    #마지막에 . @  만 남아있는거 지우기
    while True:
        if full_sent[-1] == '@' or full_sent[-1] == '.':
            full_sent = full_sent[:-1]
        else:
            break
    return full_sent

def add_whelk(token_text):    # 숫자 앞에 @ 붙이기
    for idx, word in enumerate(token_text):
        try:
            if "," in word:
                if(token_text[idx+1] == "mcg" or token_text[idx+1] == "mg" or token_text[idx+1] == "g"
                    or token_text[idx+1] == "SPU" or token_text[idx+1] == "PFU" or token_text[idx+1] == "FU" or token_text[idx+1] == "FIP"
                    and token_text[idx+2] == "@"):
                    token_text[idx] = "@" + str(word)
        except:
            pass
        
        try:
            if "." in word:
                word = float(word)
            else:
                word = int(word)
            if (token_text[idx+1] == "mcg" or token_text[idx+1] == "mg" or token_text[idx+1] == "g" 
                    or token_text[idx+1] == "SPU" or token_text[idx+1] == "PFU" or token_text[idx+1] == "FU"   
                    and token_text[idx+2] == "@"):
                token_text[idx] = "@"+str(word)
        except:
            pass
    return token_text

for i in range(0, 3):    # 7446까지
    print(i)
    
    full_sent = ''    # 잘린거 다시 합치기
    full_sent = excel_info_data[i][0]
    
    full_sent = gpt.gptTest(full_sent)
    #full_sent = clean(full_sent)    # 괄호안에 @없애기
    #full_sent = last_special_del(full_sent)    #마지막에 . @  만 남아있는거 지우기

    #sent = full_sent.replace(" ", '')    # 문자열 띄어쓰기 없게 만들기
    #kospacing_sent = spacing(full_sent)
    #token_text = okt.morphs(kospacing_sent)
    #token_text = cutting_text(token_text)    # 필요없는거 조건 추가해서 지우기
    #token_text = cutting_persent(token_text)    # @OO%@ 지우기
    #token_text = last_int_del(token_text)    # 마지막에 숫자만 있는거 지우기
    #token_text = add_whelk(token_text)

    #print()
    #print(token_text)    # 잘린 텍스트 다시 확인하기
    '''
    for idx, word in enumerate(token_text):
        if token_text[idx] == '@' or token_text[idx] == ')':
            if idx != len(token_text)-1:
                if token_text[idx+1] in stopword_list:
                    #print(token_text[idx+1])
                    token_text[idx+1] = ''
    '''
    #full_sent = ''.join(token_text)
    #print(full_sent)
    #full_sent_spacing = spacing(full_sent)
    full_sent_list.append(full_sent)
    

wb = op.load_workbook(r"etestnew2.xlsx") #Workbook 객체 생성
ws = wb["Sheet1"] #WorkSheet 객체 생성("무" Sheet)

i = 0

for text in full_sent_list:
    ws.cell(row = i+1, column = 1).value = str(excel_info_0[i])
    ws.cell(row = i+1, column = 2).value = str(excel_info_1[i])
    ws.cell(row = i+1, column = 3).value = str(excel_info_2[i])
    ws.cell(row = i+1, column = 4).value = str(excel_info_3[i])
    ws.cell(row = i+1, column = 5).value = str(excel_info_4[i])
    ws.cell(row = i+1, column = 6).value = str(excel_info_5[i])
    ws.cell(row = i+1, column = 7).value = str(text)
    ws.cell(row = i+1, column = 8).value = str(excel_info_7[i])
    i+=1
wb.save("result.xlsx") #엑셀 파일 저장(파일명 : etestnew2.xlsx)
wb.close()

print("끄으으으읕")